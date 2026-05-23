"""
HIPSTARR PROJECT 2 — Python Script 04: Output Writer
=====================================================
Pulls the complete value_flow_gap table from BigQuery
and writes two clean output files:
  1. p2_final_dataset.csv  — flat CSV for dashboard / charting
  2. p2_final_dataset.xlsx — formatted Excel workbook with
     separate sheets for full data and market summaries

Since BigQuery credentials are not available in this environment,
this script builds the output from the in-memory dataset compiled
across Scripts 01-03. All values sourced directly from prior
script outputs and confirmed BigQuery state.

Run: python3 p2_04_output_writer.py
"""

import csv
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── COMPLETE value_flow_gap DATASET ───────────────────────────────────────────
# Assembled from BigQuery confirmed values after all scripts ran.
# Columns: track_id, track_name, artist, market, track_type,
#          half_life_wks, lambda, primary_market_retention, floor_pct,
#          top_city, top_city_pct, home_market_pct, diaspora_pct,
#          blended_rate_usd, est_annual_rev_usd, catalogue_multiple,
#          est_catalogue_value_usd, vfg_score, vfg_direction, vfg_magnitude,
#          idi_score

DATA = [
    # NIGERIA
    ('NG_001','Calm Down (Remix)','Rema ft. Selena Gomez','NG','crossover',
     None,None,44.26,None,'Lagos',43.99,18.4,81.6,
     0.00085,34596.97,14,484357.58,16.56,'capture','moderate',0.1656),
    ('NG_002','Rush','Ayra Starr','NG','afrobeats',
     7.5,0.091941,47.64,26.91,'Lagos',43.99,57.97,42.03,
     0.00059,970019.38,12,11640232.56,52.173,'capture','extreme',0.5217),
    ('NG_003','Essence','Wizkid ft. Tems','NG','afrobeats',
     None,None,75.93,75.93,'Lagos',43.99,47.55,52.45,
     0.00071,693961.38,20,13879227.6,42.795,'capture','extreme',0.4279),
    ('NG_004','Soso','Omah Lay','NG','afrobeats',
     None,None,58.82,57.98,'Lagos',43.99,59.76,40.24,
     0.00055,987736.22,20,19754724.4,53.784,'capture','extreme',0.5378),
    ('NG_005','Last Last','Burna Boy','NG','afrobeats',
     None,None,26.22,38.2,'Lagos',43.99,52.62,47.38,
     0.00058,1143271.15,18,20578880.7,47.358,'capture','extreme',0.4736),
    ('NG_006','Joha','Asake','NG','afrobeats',
     None,None,48.12,61.18,'Lagos',43.99,71.94,28.06,
     0.00052,691453.81,20,13829076.2,64.746,'capture','extreme',0.6475),
    ('NG_007','Holy Ghost','Omah Lay','NG','afrobeats',
     3.4,0.202942,15.6,13.61,'Lagos',43.99,64.67,35.33,
     0.00051,3562394.0,8,28499152.0,58.203,'capture','extreme',0.582),

    # SOUTH AFRICA
    ('ZA_001','Water','Tyla','ZA','crossover',
     None,None,None,None,'Johannesburg',31.99,31.99,68.01,
     0.00078,None,6,None,19.9937,'capture','moderate',0.1999),
    ('ZA_002','Mnike','Tyler ICU ft. DJ Maphorisa et al.','ZA','amapiano',
     14.5,0.047685,36.13,3.04,'Johannesburg',55.53,55.53,44.47,
     0.00032,536699.43,6,3220196.58,34.7062,'capture','high',0.3471),
    ('ZA_003','Tshwala Bam','TitoM & Yuppe ft. SNE','ZA','amapiano',
     6.1,0.11397,12.37,9.54,'Johannesburg',19.72,19.72,80.28,
     0.00025,547882.35,8,4383058.8,12.325,'capture','moderate',0.1232),
    ('ZA_004','Asibe Happy','Kabza De Small ft. Ami Faku','ZA','amapiano',
     3.6,0.19007,20.57,23.51,'Johannesburg',76.05,76.05,23.95,
     0.00035,188779.44,12,2265353.28,47.5313,'capture','extreme',0.4753),
    ('ZA_005','Abo Mvelo','Daliwonga ft. Mellow & Sleazy','ZA','amapiano',
     7.7,0.09019,27.16,17.31,'Johannesburg',35.78,35.78,64.22,
     0.00019,97413.73,12,1168964.76,22.3625,'capture','moderate',0.2236),
    ('ZA_006','Ghost','Kamo Mphela ft. Daliwonga & Felo Le Tee','ZA','amapiano',
     None,None,None,None,'Johannesburg',24.91,24.91,75.09,
     0.00013,None,6,None,15.5687,'capture','moderate',0.1557),
    ('ZA_007','Yahyuppiyah','Uncle Waffles ft. Tony Duardo et al.','ZA','amapiano',
     10.1,0.068602,20.15,2.34,'Johannesburg',27.67,27.67,72.33,
     0.00018,128432.89,6,770597.34,17.2938,'capture','moderate',0.1729),

    # BRAZIL
    ('BR_001','Envolver','Anitta','BR','latin_pop',
     0.7,1.023676,16.25,22.99,'São Paulo',25.17,25.17,74.83,
     0.00097,156818.78,12,1881825.36,13.8435,'capture','moderate',0.1384),
    ('BR_002','Funk Rave','Anitta','BR','latin_pop',
     0.9,0.753218,23.69,22.52,'São Paulo',59.63,59.63,40.37,
     0.00123,136955.29,12,1643463.48,32.7965,'capture','high',0.328),
    ('BR_003','Nosso Quadro','AgroPlay ft. Ana Castela','BR','sertanejo',
     22.3,0.031062,47.1,0.0,'São Paulo',55.24,55.24,44.76,
     0.00111,1225978.09,6,7355868.54,30.382,'capture','high',0.3038),
    ('BR_004','Não Que Eu Vá','Os Barões da Pisadinha','BR','forró',
     None,None,None,None,'São Paulo',74.59,74.59,25.41,
     0.00149,None,6,None,41.0245,'capture','extreme',0.4102),
    ('BR_005','Modo Turbo','Luisa Sonza ft. Pabllo Vittar & Anitta','BR','latin_pop',
     5.3,0.130138,16.4,15.67,'São Paulo',70.57,70.57,29.43,
     0.00141,360004.63,12,4320055.56,38.8135,'capture','high',0.3881),
    ('BR_006','Let\'s Go 4','DJ GBR ft. MC GH do 7 et al.','BR','funk',
     22.2,0.031191,40.83,0.0,'São Paulo',76.55,76.55,23.45,
     0.00153,830734.67,6,4984408.02,42.1025,'capture','extreme',0.421),
    ('BR_007','Boiadeira','Ana Castela','BR','sertanejo',
     None,None,None,None,'São Paulo',72.3,72.3,27.7,
     0.00145,None,6,None,39.765,'capture','high',0.3977),

    # MEXICO
    ('MX_001','Ella Baila Sola','Eslabon Armado ft. Peso Pluma','MX','corrido',
     2.2,0.312628,59.17,45.81,'Mexico City',21.99,21.99,78.01,
     0.00163,5374671.53,18,96744087.54,9.8955,'low_capture','low',0.099),
    ('MX_002','La Bebé (Remix)','Peso Pluma ft. Yng Lvcas et al.','MX','corrido',
     7.2,0.096113,19.11,12.08,'Mexico City',28.34,28.34,71.66,
     0.0011,1874059.43,8,14992475.44,12.753,'capture','moderate',0.1275),
    ('MX_003','No Se Va','Grupo Frontera','MX','norteño',
     None,None,100.0,None,'Mexico City',20.04,20.04,79.96,
     0.00158,380856.62,14,5331992.68,9.018,'low_capture','low',0.0902),
    ('MX_004','Diamantes','Natanael Cano','MX','corrido',
     1.6,0.433645,17.14,59.19,'Mexico City',38.55,38.55,61.45,
     0.00157,309834.59,20,6196691.8,17.3475,'capture','moderate',0.1735),
    ('MX_005','Que Onda','Calle 24 ft. Chino Pacas & Fuerza Regida','MX','corrido',
     15.3,0.045224,42.34,17.9,'Mexico City',24.07,24.07,75.93,
     0.00158,3464893.13,12,41578717.56,10.8315,'capture','moderate',0.1083),
    ('MX_006','El Azul','Junior H ft. Peso Pluma','MX','corrido',
     11.9,0.058431,32.62,12.61,'Mexico City',29.62,29.62,70.38,
     0.00151,4342225.46,8,34737803.68,13.329,'capture','moderate',0.1333),
    ('MX_007','El Belicón','Peso Pluma ft. Raul Vega','MX','corrido',
     11.6,0.059748,38.65,0.0,'Mexico City',37.2,37.2,62.8,
     0.00147,229078.03,6,1374468.18,16.74,'capture','moderate',0.1674),

    # INDIA
    ('IN_001','Kesariya','Arijit Singh ft. Pritam','IN','bollywood',
     1.6,0.422998,75.07,70.59,'Mumbai',52.69,52.69,47.31,
     0.00021,546087.89,20,10921757.8,44.7865,'capture','extreme',0.4479),
    ('IN_002','Chaleya','Arijit Singh ft. Anirudh Ravichander & Shilpa Rao','IN','bollywood',
     None,None,None,None,'Mumbai',44.45,44.45,55.55,
     0.00018,None,6,None,37.7825,'capture','high',0.3778),
    ('IN_003','Softly','Karan Aujla ft. Ikky','IN','punjabi',
     6.2,0.112662,40.66,41.66,'Mumbai',51.1,51.1,48.9,
     0.0002,574296.34,18,10337334.12,43.435,'capture','extreme',0.4344),
    ('IN_004','Pehle Bhi Main','Vishal Mishra ft. Raj Shekhar','IN','bollywood',
     9.1,0.076227,28.26,15.41,'Mumbai',53.45,53.45,46.55,
     0.00021,3749867.68,12,44998412.16,45.4325,'capture','extreme',0.4543),
    ('IN_005','Husn','Anuv Jain','IN','indie',
     9.3,0.074912,33.42,26.66,'Mumbai',50.75,50.75,49.25,
     0.00021,2391186.59,12,28694239.08,43.1375,'capture','extreme',0.4314),
    ('IN_006','Satranga','Arijit Singh ft. Shreyas Puranik','IN','bollywood',
     8.7,0.079237,39.86,20.82,'Mumbai',50.33,50.33,49.67,
     0.0002,3154985.27,12,37859823.24,42.7805,'capture','extreme',0.4278),
    ('IN_007','Tere Vaaste','Varun Jain ft. Alamash Faridi et al.','IN','bollywood',
     7.8,0.089016,19.33,10.22,'Mumbai',57.5,57.5,42.5,
     0.00023,2856148.01,8,22849184.08,48.875,'capture','extreme',0.4887),

    # SAUDI ARABIA
    ('SA_001','Lammah','Ayed','SA','khaleeji',
     13.9,0.050003,43.94,6.99,'Riyadh',53.26,53.26,46.74,
     0.00247,165277.3,8,1322218.4,19.9725,'capture','moderate',0.1997),
    ('SA_002','Alam Aloshag','Khaled Almuthafar ft. Ayed','SA','khaleeji',
     24.3,0.028491,53.84,0.0,'Riyadh',45.4,45.4,54.6,
     0.00202,42660.74,6,255964.44,17.025,'capture','moderate',0.1702),
    ('SA_003','Ya Ibn Al Awadim','Abdul Majeed Abdullah','SA','khaleeji',
     7.0,0.098918,40.73,37.18,'Riyadh',10.1,10.1,89.9,
     0.00046,68918.93,16,1102702.88,3.7875,'low_capture','low',0.0379),
    ('SA_004','Mshtaglk','Mohamed AlSalim ft. Helly Luv','SA','khaleeji',
     6.7,0.103796,27.51,23.3,'Riyadh',2.52,2.52,97.48,
     0.00056,65395.5,12,784746.0,0.945,'low_capture','low',0.0095),
    ('SA_005','Da Elly 7sal','Balqees','SA','khaleeji',
     9.3,0.074141,56.71,36.28,'Riyadh',23.33,23.33,76.67,
     0.00166,148918.75,16,2382700.0,8.7487,'low_capture','low',0.0875),
    ('SA_006','Ela Yakbar Hazi','Abdullah Al Farwan','SA','khaleeji',
     1.4,0.480851,31.16,32.98,'Riyadh',10.1,10.1,89.9,
     0.00046,50058.17,16,800930.72,3.7875,'low_capture','low',0.0379),
    ('SA_007','Kel Ahebek','Fouad Abdulwahed','SA','khaleeji',
     5.8,0.118951,51.87,45.84,'Riyadh',23.89,23.89,76.11,
     0.00119,63044.07,18,1134793.26,8.9588,'low_capture','low',0.0896),
]

COLUMNS = [
    'track_id','track_name','artist','market','track_type',
    'half_life_wks','lambda','primary_market_retention','floor_pct',
    'top_city','top_city_pct','home_market_pct','diaspora_pct',
    'blended_rate_usd','est_annual_rev_usd','catalogue_multiple',
    'est_catalogue_value_usd','vfg_score','vfg_direction','vfg_magnitude',
    'idi_score'
]

GOLD   = 'D4AF37'
GREEN  = '34D399'
WHITE  = 'F0EEF8'
DARK   = '0d0d0d'
DARKER = '0a0a0a'
HEADER = '1a1a2e'

MARKET_COLORS = {
    'NG': 'FF6B35',  # orange
    'ZA': '34D399',  # green
    'BR': 'FFD700',  # gold
    'MX': 'C084FC',  # purple
    'IN': '60A5FA',  # blue
    'SA': 'F87171',  # red
}


def fmt(val, decimals=2):
    if val is None:
        return ''
    if isinstance(val, float):
        return round(val, decimals)
    return val


def write_csv(rows, path):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(dict(zip(COLUMNS, row)))
    print(f"CSV written: {path}")


def write_xlsx(rows, path):
    wb = Workbook()

    # ── Sheet 1: Full Dataset ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Full Dataset'

    # Title block
    ws['A1'] = 'HIPSTARR PROJECT 2 — AFROBEATS VIRALITY VS LONGEVITY'
    ws['A1'].font = Font(bold=True, color=GOLD, name='Arial', size=13)
    ws['A2'] = 'Value Flow Gap Analysis | 42 Tracks | 6 Markets'
    ws['A2'].font = Font(bold=True, color='888888', name='Arial', size=10)
    ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d")} | Hipstarr Music Research'
    ws['A3'].font = Font(color='555555', name='Arial', size=9)

    # Headers
    header_fill = PatternFill('solid', start_color=HEADER)
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=5, column=col_idx, value=col_name.upper().replace('_', ' '))
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for row_idx, row in enumerate(rows, 6):
        market = row[3]
        mkt_color = MARKET_COLORS.get(market, '333333')
        row_fill = PatternFill('solid', start_color=DARKER)

        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=fmt(val))
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='left')

            # Style track_id with market colour
            if col_idx == 1:
                cell.font = Font(bold=True, color=mkt_color, name='Arial', size=9)
            elif col_idx in (18, 19, 20):  # VFG columns
                cell.font = Font(color=GREEN, name='Arial', size=9)
            elif col_idx == 21:  # IDI
                cell.font = Font(color=GOLD, name='Arial', size=9)
            else:
                cell.font = Font(color='CCCCCC', name='Arial', size=9)

            # Number formats
            if col_idx in (14,):  # blended_rate
                cell.number_format = '$#,##0.00000'
            elif col_idx in (15, 17):  # revenue / catalogue value
                cell.number_format = '$#,##0'
            elif col_idx in (11, 12, 13, 7, 8, 9, 18, 21):
                cell.number_format = '0.00'

    # Column widths
    col_widths = [10,28,35,7,12,10,10,12,8,14,9,9,9,12,14,8,16,9,10,10,8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = 'B6'

    # ── Sheet 2: Market Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet('Market Summary')
    ws2['A1'] = 'MARKET SUMMARY'
    ws2['A1'].font = Font(bold=True, color=GOLD, name='Arial', size=12)

    mkt_headers = ['Market','Tracks','Avg VFG','Avg IDI','Avg Home%',
                   'Avg Diaspora%','Avg Blended Rate','Portfolio Value $M',
                   'Avg Multiple','Market IDI']
    for col_idx, h in enumerate(mkt_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
        cell.fill = PatternFill('solid', start_color=HEADER)
        cell.alignment = Alignment(horizontal='center')

    market_idi = {'NG':0.9,'ZA':0.625,'BR':0.55,'MX':0.45,'IN':0.85,'SA':0.375}
    markets = ['NG','ZA','BR','MX','IN','SA']

    for row_idx, mkt in enumerate(markets, 4):
        mkt_rows = [r for r in rows if r[3] == mkt]
        n = len(mkt_rows)
        avg_vfg  = round(sum(r[17] for r in mkt_rows)/n, 2)
        avg_idi  = round(sum(r[20] for r in mkt_rows)/n, 4)
        avg_home = round(sum(r[11] for r in mkt_rows)/n, 1)
        avg_dias = round(sum(r[12] for r in mkt_rows)/n, 1)
        avg_rate = round(sum(r[13] for r in mkt_rows)/n, 5)
        total_val = sum(r[16] for r in mkt_rows if r[16])
        avg_mult  = round(sum(r[15] for r in mkt_rows)/n, 1)
        m_idi    = market_idi[mkt]

        row_data = [mkt, n, avg_vfg, avg_idi, avg_home, avg_dias,
                    avg_rate, round(total_val/1e6, 2), avg_mult, m_idi]
        mkt_color = MARKET_COLORS.get(mkt, '333333')
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PatternFill('solid', start_color=DARKER)
            cell.alignment = Alignment(horizontal='center')
            if col_idx == 1:
                cell.font = Font(bold=True, color=mkt_color, name='Arial', size=10)
            else:
                cell.font = Font(color='CCCCCC', name='Arial', size=9)

    # Portfolio total row
    total_row = len(markets) + 4
    valued = [r for r in rows if r[16]]
    total_val = sum(r[16] for r in valued)
    ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(
        bold=True, color=GOLD, name='Arial', size=10)
    ws2.cell(row=total_row, column=8, value=round(total_val/1e6, 2)).font = Font(
        bold=True, color=GOLD, name='Arial', size=10)

    for i, w in enumerate([10,8,10,10,10,12,14,16,10,10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Top Tracks by VFG ───────────────────────────────────────────
    ws3 = wb.create_sheet('VFG Ranking')
    ws3['A1'] = 'TRACKS RANKED BY VFG SCORE'
    ws3['A1'].font = Font(bold=True, color=GOLD, name='Arial', size=12)
    ws3['A2'] = 'Higher score = larger capture gap (more value locked in low-rate home market)'
    ws3['A2'].font = Font(color='888888', name='Arial', size=9)

    vfg_headers = ['Rank','Track ID','Track Name','Market','Track Type',
                   'VFG Score','VFG Direction','VFG Magnitude',
                   'Home%','Diaspora%','IDI Score','Cat Value $M']
    for col_idx, h in enumerate(vfg_headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
        cell.fill = PatternFill('solid', start_color=HEADER)

    sorted_rows = sorted(rows, key=lambda x: -x[17])
    for rank, row in enumerate(sorted_rows, 1):
        r_idx = rank + 4
        mkt_color = MARKET_COLORS.get(row[3], '333333')
        cat_val_m = round(row[16]/1e6, 2) if row[16] else None
        row_data = [rank, row[0], row[1], row[3], row[4],
                    fmt(row[17]), row[18], row[19],
                    fmt(row[11]), fmt(row[12]), fmt(row[20]), cat_val_m]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r_idx, column=col_idx, value=val)
            cell.fill = PatternFill('solid', start_color=DARKER)
            if col_idx == 2:
                cell.font = Font(bold=True, color=mkt_color, name='Arial', size=9)
            elif col_idx == 6:
                cell.font = Font(bold=True, color=GREEN, name='Arial', size=9)
            else:
                cell.font = Font(color='CCCCCC', name='Arial', size=9)

    for i, w in enumerate([6,10,30,8,12,10,12,12,8,9,9,12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"Excel written: {path}")


def main():
    rows = DATA
    print(f"Processing {len(rows)} tracks...")

    csv_path  = '/mnt/user-data/outputs/p2_final_dataset.csv'
    xlsx_path = '/mnt/user-data/outputs/p2_final_dataset.xlsx'

    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)

    # Summary stats
    valued = [r for r in rows if r[16]]
    total_portfolio = sum(r[16] for r in valued)
    print(f"\nSummary:")
    print(f"  Total tracks: {len(rows)}")
    print(f"  Tracks with catalogue value: {len(valued)}")
    print(f"  Portfolio total: ${total_portfolio:,.0f}")
    print(f"  Avg VFG score: {round(sum(r[17] for r in rows)/len(rows), 2)}")
    print(f"  Avg IDI score: {round(sum(r[20] for r in rows)/len(rows), 4)}")
    print("\nFiles ready.")


if __name__ == '__main__':
    main()
