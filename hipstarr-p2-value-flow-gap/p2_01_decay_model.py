"""
HIPSTARR PROJECT 2 — Python Script 01: Decay Model
====================================================
Reads Kworb weekly stream data from p2_kworb_data.xlsx.
For each track:
  1. Extracts primary market weekly time series
  2. Fits exponential decay model where appropriate
  3. Calculates: half_life_wks, lambda, primary_market_retention,
     global_retention, floor_pct, streams_6m, streams_12m
  4. Handles special cases: non-decay, two-lifecycle, Tier 3, Ramadan
  5. Prints results table (BigQuery UPDATE statements generated separately)

Run from terminal:
  python3 p2_01_decay_model.py

Requirements:
  pip install openpyxl pandas numpy scipy
"""

import re
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from scipy.optimize import curve_fit
from openpyxl import load_workbook

# ── CONFIG ────────────────────────────────────────────────────────────────────

EXCEL_PATH = '/mnt/user-data/outputs/p2_kworb_data.xlsx'
DATA_START_ROW = 10   # Row index (0-based) where weekly data begins in each sheet

# Primary market column per track
PRIMARY_MARKET = {
    'NG_001': 'NG', 'NG_002': 'NG', 'NG_003': 'NG', 'NG_004': 'NG',
    'NG_005': 'NG', 'NG_006': 'NG', 'NG_007': 'NG',
    'ZA_001': 'ZA', 'ZA_002': 'ZA', 'ZA_003': 'ZA', 'ZA_004': 'ZA',
    'ZA_005': 'ZA', 'ZA_006': 'ZA', 'ZA_007': 'ZA',
    'BR_001': 'BR', 'BR_002': 'BR', 'BR_003': 'BR', 'BR_004': 'BR',
    'BR_005': 'BR', 'BR_006': 'BR', 'BR_007': 'BR',
    'MX_001': 'MX', 'MX_002': 'MX', 'MX_003': 'MX', 'MX_004': 'MX',
    'MX_005': 'MX', 'MX_006': 'MX', 'MX_007': 'MX',
    'IN_001': 'IN', 'IN_002': 'IN', 'IN_003': 'IN', 'IN_004': 'IN',
    'IN_005': 'IN', 'IN_006': 'IN', 'IN_007': 'IN',
    'SA_001': 'SA', 'SA_002': 'SA', 'SA_003': 'SA', 'SA_004': 'SA',
    'SA_005': 'SA', 'SA_006': 'SA', 'SA_007': 'SA',
}

# Sheet name mapping (handles duplicates - use first occurrence)
SHEET_MAP = {
    'NG_001': 'NG_001 Calm Down Remix',
    'NG_002': 'NG_002 Rush',
    'NG_003': 'NG_003 Essence',
    'NG_004': 'NG_004 Soso',
    'NG_005': 'NG_005 Last Last',
    'NG_006': 'NG_006 Joha',
    'NG_007': 'NG_007 Holy Ghost',
    'ZA_001': 'ZA_001 Water',
    'ZA_002': 'ZA_002 Mnike',
    'ZA_003': 'ZA_003 Tshwala Bam',
    'ZA_004': 'ZA_004 Asibe Happy',
    'ZA_005': 'ZA_005 Abo Mvelo',
    'ZA_006': 'ZA_006 Ghost',
    'ZA_007': 'ZA_007 Yahyuppiyah',
    'BR_001': 'BR_001 Envolver',
    'BR_002': 'BR_002 Funk Rave',
    'BR_003': 'BR_003 Nosso Quadro',
    'BR_004': 'BR_004 Nao Que Eu Va',
    'BR_005': 'BR_005 Modo Turbo',
    'BR_006': 'BR_006 Lets Go 4',
    'BR_007': 'BR_007 Boiadeira',
    'MX_001': 'MX_001 Ella Baila Sola',
    'MX_002': 'MX_002 La Bebe Remix',
    'MX_003': 'MX_003 No Se Va',
    'MX_004': 'MX_004 Diamantes',
    'MX_005': 'MX_005 Que Onda',
    'MX_006': 'MX_006 El Azul',
    'MX_007': 'MX_007 El Belicon',
    'IN_001': 'IN_001 Kesariya',
    'IN_002': 'IN_002 Chaleya',
    'IN_003': 'IN_003 Softly',
    'IN_004': 'IN_004 Pehle Bhi Main',
    'IN_005': 'IN_005 Husn',
    'IN_006': 'IN_006 Satranga',
    'IN_007': 'IN_007 Tere Vaaste',
    'SA_001': 'SA_001 Lammah',
    'SA_002': 'SA_002 Alam Aloshag',
    'SA_003': 'SA_003 Ya Ibn Al Awadim',
    'SA_004': 'SA_004 Mshtaglk',
    'SA_005': 'SA_005 Da Elly 7sal',
    'SA_006': 'SA_006 Ela Yakbar Hazi',
    'SA_007': 'SA_007 Kel Ahebek',
}

# Special case flags (from Section 12.8 of methodology notes)
TIER_3 = {'ZA_001', 'ZA_006', 'BR_004', 'BR_007', 'IN_002'}
TIER_2 = {'IN_001', 'MX_003'}
NON_DECAY = {'NG_003', 'NG_004', 'NG_006', 'NG_005'}
TWO_LIFECYCLE = {'NG_007', 'MX_004', 'MX_005', 'BR_002'}
RAMADAN_AFFECTED = {'SA_003', 'SA_005', 'SA_007'}

# Ramadan exclusion windows (approximate — March/April weeks)
RAMADAN_WINDOWS = [
    ('2023-03-22', '2023-04-21'),
    ('2024-03-10', '2024-04-09'),
]

# Two-lifecycle cutoff dates (use data only up to this date for decay fit)
LIFECYCLE_CUTOFF = {
    'NG_007': '2025-03-01',   # Before Apr 2026 re-peak
    'MX_004': '2021-09-01',   # First lifecycle only (2021)
    'MX_005': '2024-10-01',   # Before Apr 2026 re-surge
    'BR_002': '2023-08-01',   # First chart period only
}


# ── DATA LOADING ──────────────────────────────────────────────────────────────

def load_track_series(wb, track_id):
    """
    Load weekly primary market stream series for a track.
    Returns: dict with keys 'dates', 'streams', 'peak', 'columns', 'all_data'
    """
    sheet_name = SHEET_MAP[track_id]
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row (row containing 'Date')
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if row[0] == 'Date':
            header_row_idx = i
            break

    if header_row_idx is None:
        return None

    headers = [str(h) if h is not None else '' for h in all_rows[header_row_idx]]

    # Get primary market column index
    primary_col = PRIMARY_MARKET[track_id]
    if primary_col not in headers:
        return None
    col_idx = headers.index(primary_col)

    # Parse weekly rows (data starts 3 rows after header: Total, Peak, then data)
    data_start = header_row_idx + 3
    dates = []
    streams = []
    peak_val = None

    # Get peak from Peak row
    peak_row = all_rows[header_row_idx + 2]
    try:
        peak_val = int(peak_row[col_idx]) if isinstance(peak_row[col_idx], (int, float)) else None
    except:
        peak_val = None

    for row in all_rows[data_start:]:
        date_val = row[0]
        stream_val = row[col_idx] if col_idx < len(row) else None

        # Parse date
        if date_val is None:
            continue
        try:
            if isinstance(date_val, str) and '/' in date_val:
                dt = datetime.strptime(date_val, '%Y/%m/%d')
            elif isinstance(date_val, datetime):
                dt = date_val
            else:
                continue
        except:
            continue

        # Parse stream value
        if stream_val is None or stream_val == '--' or stream_val == '':
            continue
        try:
            val = int(stream_val)
        except:
            continue

        dates.append(dt)
        streams.append(val)

    return {
        'dates': dates,
        'streams': streams,
        'peak': peak_val,
        'primary_col': primary_col,
    }


# ── RAMADAN FILTER ────────────────────────────────────────────────────────────

def filter_ramadan(dates, streams):
    """Remove Ramadan-period weeks from Saudi tracks before fitting."""
    filtered_dates, filtered_streams = [], []
    for dt, s in zip(dates, streams):
        in_ramadan = False
        for start_str, end_str in RAMADAN_WINDOWS:
            start = datetime.strptime(start_str, '%Y-%m-%d')
            end = datetime.strptime(end_str, '%Y-%m-%d')
            if start <= dt <= end:
                in_ramadan = True
                break
        if not in_ramadan:
            filtered_dates.append(dt)
            filtered_streams.append(s)
    return filtered_dates, filtered_streams


# ── LIFECYCLE CUTOFF ──────────────────────────────────────────────────────────

def apply_cutoff(track_id, dates, streams):
    """Trim to first lifecycle for two-lifecycle tracks."""
    if track_id not in LIFECYCLE_CUTOFF:
        return dates, streams
    cutoff = datetime.strptime(LIFECYCLE_CUTOFF[track_id], '%Y-%m-%d')
    d_out, s_out = [], []
    for dt, s in zip(dates, streams):
        if dt <= cutoff:
            d_out.append(dt)
            s_out.append(s)
    return d_out, s_out


# ── DECAY MODEL ───────────────────────────────────────────────────────────────

def exp_decay(t, lam, floor_frac):
    """S(t) = peak × ((1-floor_frac) × e^(-λt) + floor_frac)"""
    return (1 - floor_frac) * np.exp(-lam * t) + floor_frac


def fit_decay(dates, streams, peak):
    """
    Fit exponential decay to normalised weekly stream series.
    Returns: (lambda, floor_frac, r_squared) or None if fit fails.
    """
    if len(streams) < 4 or peak is None or peak == 0:
        return None

    # Find peak position and use only post-peak data
    arr = np.array(streams)
    peak_idx = np.argmax(arr)
    post_dates = dates[peak_idx:]
    post_streams = arr[peak_idx:]

    if len(post_streams) < 3:
        return None

    # Normalise: S_norm(t) = streams / peak
    peak_actual = post_streams[0]
    if peak_actual == 0:
        return None
    s_norm = post_streams / peak_actual

    # Time in weeks from peak
    t_weeks = np.array([(d - post_dates[0]).days / 7.0 for d in post_dates])

    try:
        # Initial guess: half-life ~26 weeks, floor ~0.05
        lam0 = np.log(2) / 26
        popt, pcov = curve_fit(
            exp_decay,
            t_weeks,
            s_norm,
            p0=[lam0, 0.05],
            bounds=([0, 0], [2.0, 0.95]),
            maxfev=5000
        )
        lam, floor_frac = popt

        # R-squared
        s_pred = exp_decay(t_weeks, lam, floor_frac)
        ss_res = np.sum((s_norm - s_pred) ** 2)
        ss_tot = np.sum((s_norm - np.mean(s_norm)) ** 2)
        r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0

        return lam, floor_frac, round(r2, 4)

    except Exception:
        return None


# ── STREAM SUMS ───────────────────────────────────────────────────────────────

def compute_stream_sums(dates, streams, peak_date):
    """Sum streams within 6-month and 12-month windows from peak date."""
    six_m = 0
    twelve_m = 0
    for dt, s in zip(dates, streams):
        days = (dt - peak_date).days
        if 0 <= days <= 182:
            six_m += s
        if 0 <= days <= 365:
            twelve_m += s
    return six_m, twelve_m


# ── RETENTION CALCULATION ─────────────────────────────────────────────────────

def compute_retention(dates, streams, peak, weeks=26):
    """
    Estimate retention at N weeks post-peak.
    Uses nearest actual data point to the target date.
    """
    if not dates or peak is None or peak == 0:
        return None

    arr = np.array(streams)
    peak_idx = np.argmax(arr)
    peak_date = dates[peak_idx]
    target = peak_date + timedelta(weeks=weeks)

    # Find closest data point to target
    min_diff = None
    closest_val = None
    for dt, s in zip(dates, streams):
        diff = abs((dt - target).days)
        if min_diff is None or diff < min_diff:
            min_diff = diff
            closest_val = s

    if closest_val is None:
        return None

    return round(closest_val / peak * 100, 2)


# ── MAIN PROCESSING ───────────────────────────────────────────────────────────

def process_track(wb, track_id):
    """Process a single track and return metrics dict."""
    result = {
        'track_id': track_id,
        'half_life_wks': None,
        'lambda': None,
        'r_squared': None,
        'primary_market_retention': None,
        'global_retention': None,
        'floor_pct': None,
        'streams_6m': None,
        'streams_12m': None,
        'model_type': None,
        'notes': '',
    }

    # Tier 3: no decay modelling
    if track_id in TIER_3:
        result['model_type'] = 'tier3_skip'
        result['notes'] = 'Insufficient Kworb data for decay modelling'
        return result

    data = load_track_series(wb, track_id)
    if not data or len(data['streams']) < 3:
        result['model_type'] = 'insufficient_data'
        result['notes'] = 'Fewer than 3 data points'
        return result

    dates = data['dates']
    streams = data['streams']
    peak = data['peak']

    # Apply filters
    if track_id in RAMADAN_AFFECTED:
        dates, streams = filter_ramadan(dates, streams)
        result['notes'] += 'Ramadan weeks excluded. '

    if track_id in TWO_LIFECYCLE:
        dates, streams = apply_cutoff(track_id, dates, streams)
        result['notes'] += f"Lifecycle 1 only (cutoff {LIFECYCLE_CUTOFF[track_id]}). "

    if len(streams) < 3:
        result['model_type'] = 'insufficient_after_filter'
        return result

    # Find peak position
    arr = np.array(streams)
    peak_idx = np.argmax(arr)
    peak_actual = arr[peak_idx]
    peak_date = dates[peak_idx]

    if peak is None:
        peak = int(peak_actual)

    # Non-decay tracks: compute retention and floor without fitting
    if track_id in NON_DECAY:
        result['model_type'] = 'non_decay'
        result['notes'] += 'Non-decay track — domestic base does not decay.'
        # Use last available value as floor
        last_val = streams[-1]
        result['floor_pct'] = round(last_val / peak * 100, 2)
        result['primary_market_retention'] = compute_retention(
            dates, streams, peak, weeks=26)
        s6, s12 = compute_stream_sums(dates, streams, peak_date)
        result['streams_6m'] = s6 if s6 > 0 else None
        result['streams_12m'] = s12 if s12 > 0 else None
        return result

    # Fit decay model
    fit = fit_decay(dates, streams, peak)

    if fit is not None:
        lam, floor_frac, r2 = fit
        half_life = round(np.log(2) / lam, 1) if lam > 0 else None
        result['lambda'] = round(lam, 6)
        result['half_life_wks'] = half_life
        result['floor_pct'] = round(floor_frac * 100, 2)
        result['r_squared'] = r2
        result['model_type'] = 'exponential_decay'
        if r2 < 0.5:
            result['notes'] += f'Low R² ({r2}) — fit unreliable. '
    else:
        result['model_type'] = 'fit_failed'
        result['notes'] += 'Curve fit failed — using empirical values only. '

    # Retention and stream sums
    result['primary_market_retention'] = compute_retention(
        dates, streams, peak, weeks=26)
    s6, s12 = compute_stream_sums(dates, streams, peak_date)
    result['streams_6m'] = s6 if s6 > 0 else None
    result['streams_12m'] = s12 if s12 > 0 else None

    return result


# ── RUN ───────────────────────────────────────────────────────────────────────

def main():
    print("Loading Kworb Excel file...")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    track_ids = list(SHEET_MAP.keys())
    results = []

    for track_id in track_ids:
        r = process_track(wb, track_id)
        results.append(r)

    wb.close()

    # Print results table
    print("\n" + "="*110)
    print(f"{'Track':<10} {'Model':<20} {'λ':<10} {'Half-life':<12} {'Ret-26wk':<12} {'Floor%':<10} {'R²':<8} {'6M':<14} {'12M':<14}")
    print("="*110)
    for r in results:
        print(
            f"{r['track_id']:<10} "
            f"{str(r['model_type']):<20} "
            f"{str(r['lambda']):<10} "
            f"{str(r['half_life_wks']):<12} "
            f"{str(r['primary_market_retention']):<12} "
            f"{str(r['floor_pct']):<10} "
            f"{str(r['r_squared']):<8} "
            f"{str(r['streams_6m']):<14} "
            f"{str(r['streams_12m']):<14} "
        )

    print("\n\nFLAGS AND NOTES:")
    print("-"*80)
    for r in results:
        if r['notes']:
            print(f"  {r['track_id']}: {r['notes']}")

    # Generate BigQuery UPDATE statements
    print("\n\n" + "="*80)
    print("-- BigQuery UPDATE statements for value_flow_gap")
    print("="*80)
    for r in results:
        hl = f"{r['half_life_wks']}" if r['half_life_wks'] is not None else 'NULL'
        lam = f"{r['lambda']}" if r['lambda'] is not None else 'NULL'
        ret = f"{r['primary_market_retention']}" if r['primary_market_retention'] is not None else 'NULL'
        flr = f"{r['floor_pct']}" if r['floor_pct'] is not None else 'NULL'
        print(
            f"UPDATE `hipstarr_p2.value_flow_gap` "
            f"SET half_life_wks={hl}, lambda={lam}, "
            f"primary_market_retention={ret}, floor_pct={flr} "
            f"WHERE track_id='{r['track_id']}';"
        )

    print("\n-- BigQuery UPDATE statements for market_streams streams_6m/12m")
    print("-"*80)
    for r in results:
        s6 = f"{r['streams_6m']}" if r['streams_6m'] is not None else 'NULL'
        s12 = f"{r['streams_12m']}" if r['streams_12m'] is not None else 'NULL'
        print(
            f"UPDATE `hipstarr_p2.market_streams` "
            f"SET streams_6m={s6}, streams_12m={s12} "
            f"WHERE track_id='{r['track_id']}' AND is_primary_market=TRUE;"
        )

    print("\nDone. Review output above before running BigQuery statements.")
    return results


if __name__ == '__main__':
    main()
